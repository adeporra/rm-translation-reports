#!/usr/bin/env python3
"""
AEM PROD Author Translation Log Analyzer
=========================================
Parses aemerror log files from AEM Cloud Service PROD author,
extracts translation job lifecycle data, errors, retries, and metrics,
and generates consolidated CSV, HTML, and XLSX reports.

Usage:
    python analyze_translation_log.py <log_file> [--reports-dir <dir>]
"""

import sys
import os
import re
import csv
import json
import subprocess
import argparse
import tempfile
from datetime import datetime
from collections import defaultdict, OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CSV_COLUMNS = [
    'Date', '#', 'Job ID', 'Project', 'Language', 'Source Language',
    'Connector', 'Method', 'Status', 'Created', 'Content Add Start',
    'Content Add End', 'InProgress', 'ReadyForReview', 'Approved/Error',
    'Content Gathering (ms)', 'Content Gathering',
    'Created→InProg (ms)', 'Created→InProg',
    'InProg→Ready (ms)', 'InProg→Ready',
    'Ready→Approved (ms)', 'Ready→Approved',
    'Total Duration (ms)', 'Total Duration',
    'Content Objects', 'Content Size', 'Errors', 'Retries', 'Retry Detail',
]

EVENT_NAMES = [
    'TranslationJobCreated',
    'TranslationJobContentAdditionStarted',
    'TranslationJobContentAdditionCompleted',
    'TranslationJobInProgress',
    'TranslationJobReadyForReview',
    'TranslationJobApproved',
    'TranslationJobError',
]

ERROR_CATEGORIES = OrderedDict([
    ('Language copy not found', 'Language copy not found'),
    ('Content resource language copy creation is in unknown state', 'Content LC unknown state'),
    ('Failed to start translation', 'Failed to start translation'),
    ('Error while processing Translation project', 'Error processing Translation project'),
    ('TranslationCleanupJobConsumer', 'TranslationCleanupJobConsumer'),
    ('ThumbnailServlet', 'ThumbnailServlet'),
    ('TicketAssemblyTranslationServiceImpl', 'TicketAssemblyTranslationServiceImpl'),
    ('GCCScriptProcessor', 'GCCScriptProcessor'),
    ('ReplicationSquadListener', 'ReplicationSquadListener'),
    ('ReplicationNewsAssemblyListener', 'ReplicationNewsAssemblyListener'),
    ('ResourceUtils', 'ResourceUtils commit'),
    ('LockUtil', 'LockUtil contention'),
    ('TranslateHrefAttributes', 'TranslateHrefAttributes'),
    ('NoLocalizedResource', 'NoLocalizedResource'),
    ('TranslationStatusUpdate', 'TranslationStatusUpdate'),
    ('EmbeddedAssetReplace', 'EmbeddedAssetReplace'),
    ('Error executing workflow', 'Error executing workflow'),
])

NOISE_PATTERNS = [
    'PublishLiveSummaryTranslateStep',
    'El idioma del CF no es en-us',
    'ReplicationCamerasListener',
]

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def parse_time(ts_str):
    """Parse HH:MM:SS.mmm -> ms from midnight."""
    try:
        parts = ts_str.split(':')
        h, m = int(parts[0]), int(parts[1])
        sp = parts[2].split('.')
        s = int(sp[0])
        ms = int(sp[1]) if len(sp) > 1 else 0
        return (h * 3600 + m * 60 + s) * 1000 + ms
    except Exception:
        return None


def ms_to_human(ms):
    """Convert ms to human string."""
    if ms is None:
        return ''
    if ms < 0:
        ms = abs(ms)
    s = ms / 1000.0
    if s < 60:
        return f'{s:.1f}s'
    mi = int(s // 60)
    rem = s % 60
    return f'{mi}m {rem:.1f}s'


def diff_ms(t1, t2):
    """Difference in ms between two HH:MM:SS.mmm strings."""
    a, b = parse_time(t1), parse_time(t2)
    if a is None or b is None:
        return None
    return b - a


def extract_date_from_filename(filename):
    """Extract YYYY-MM-DD from filenames like author_aemerror_2026-03-12.log"""
    m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(filename))
    return m.group(1) if m else None


def format_date_display(date_str):
    """2026-03-12 -> 03-12"""
    parts = date_str.split('-')
    return f'{parts[1]}-{parts[2]}' if len(parts) == 3 else date_str


def clean_project_name(raw):
    """Convert path segment to display name."""
    name = raw.replace('-', ' ').strip()
    name = re.sub(r'\s+', ' ', name)
    return name.title()

# ---------------------------------------------------------------------------
# Grep wrappers for large files
# ---------------------------------------------------------------------------

def grep_extract(logfile, pattern, outfile):
    """Extract lines matching pattern from logfile into outfile. Returns count."""
    try:
        result = subprocess.run(
            ['grep', pattern, logfile],
            capture_output=True, text=True, timeout=600
        )
        with open(outfile, 'w') as f:
            f.write(result.stdout)
        return result.stdout.count('\n')
    except Exception as e:
        print(f'  Warning: grep failed for {pattern}: {e}', file=sys.stderr)
        open(outfile, 'w').close()
        return 0


def grep_count(logfile, pattern):
    """Count lines matching pattern."""
    try:
        result = subprocess.run(
            ['grep', '-c', pattern, logfile],
            capture_output=True, text=True, timeout=300
        )
        return int(result.stdout.strip())
    except Exception:
        return 0

# ---------------------------------------------------------------------------
# Event parsing
# ---------------------------------------------------------------------------

LINE_RE = re.compile(
    r'^(\d{2}\.\d{2}\.\d{4})\s+'          # date DD.MM.YYYY
    r'(\d{2}:\d{2}:\d{2}\.\d{3})\s+'      # timestamp HH:MM:SS.mmm
    r'\[([^\]]+)\]\s+'                      # instance
    r'\*(\w+)\*\s+'                         # level
    r'\[([^\]]+)\]\s+'                      # thread
    r'(\S+)\s+'                             # class
    r'(.*)'                                 # message
)

JOB_PATH_RE = re.compile(r'jobPath:(/content/projects/([^/]+)/jcr:content/dashboard/gadgets/(translationjob\d+))')
LANG_RE = re.compile(r'jobDestinationLanguage:(\S+?)[\s,}]')
SRC_LANG_RE = re.compile(r'sourceLanguage:(\S+?)[\s,}]')
CONNECTOR_RE = re.compile(r'connectorName:(\S+?)[\s,}]')
METHOD_RE = re.compile(r'translationMethod:(\S+?)[\s,}]')
OBJ_COUNT_RE = re.compile(r'translationObjectCount:\{([^}]+)\}')
CONTENT_SIZE_RE = re.compile(r'contentSize:(\d+)')
WORKFLOW_RE = re.compile(r'JobHandler:\s*(/var/workflow/instances/[^:]+)')


def parse_events(events_file, date_str):
    """Parse translation pipeline events into job records.
    Uses (project_raw, job_id) as composite key since job IDs can repeat across projects.
    """
    jobs = {}
    thread_to_jobs = defaultdict(set)

    with open(events_file, 'r') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            m = LINE_RE.match(line)
            if not m:
                continue

            raw_date, timestamp, _instance, _level, thread, _cls, message = m.groups()

            jm = JOB_PATH_RE.search(message)
            if not jm:
                continue
            _job_path, project_raw, job_id = jm.groups()

            composite_key = f'{project_raw}/{job_id}'

            wf_match = WORKFLOW_RE.search(thread)
            workflow_prefix = wf_match.group(1) if wf_match else thread
            thread_to_jobs[workflow_prefix].add(composite_key)

            if composite_key not in jobs:
                jobs[composite_key] = {
                    'date': date_str,
                    'job_id': job_id,
                    'project': clean_project_name(project_raw),
                    'language': '',
                    'source_language': '',
                    'connector': '',
                    'method': '',
                    'status': 'UNKNOWN',
                    'created': '',
                    'content_add_start': '',
                    'content_add_end': '',
                    'in_progress': '',
                    'ready_for_review': '',
                    'approved_error': '',
                    'content_objects': '',
                    'content_size': '',
                    'errors': '',
                    'retries': '',
                    'retry_detail': '',
                    'workflow_thread': workflow_prefix,
                }

            job = jobs[composite_key]

            lm = LANG_RE.search(message)
            if lm:
                job['language'] = lm.group(1).upper()
            sm = SRC_LANG_RE.search(message)
            if sm:
                job['source_language'] = sm.group(1)
            cm = CONNECTOR_RE.search(message)
            if cm:
                job['connector'] = cm.group(1)
            mm = METHOD_RE.search(message)
            if mm:
                job['method'] = mm.group(1)

            if 'TranslationJobCreated' in message:
                job['created'] = timestamp
            elif 'TranslationJobContentAdditionStarted' in message:
                job['content_add_start'] = timestamp
                csm = CONTENT_SIZE_RE.search(message)
                if csm:
                    job['content_size'] = csm.group(1)
            elif 'TranslationJobContentAdditionCompleted' in message:
                job['content_add_end'] = timestamp
                om = OBJ_COUNT_RE.search(message)
                if om:
                    job['content_objects'] = om.group(1)
            elif 'TranslationJobInProgress' in message:
                job['in_progress'] = timestamp
                job['status'] = 'IN_PROGRESS'
                if not job['content_objects']:
                    om = OBJ_COUNT_RE.search(message)
                    if om:
                        job['content_objects'] = om.group(1)
            elif 'TranslationJobReadyForReview' in message:
                job['ready_for_review'] = timestamp
                job['status'] = 'READY_FOR_REVIEW'
            elif 'TranslationJobApproved' in message:
                job['approved_error'] = timestamp
                job['status'] = 'APPROVED'
            elif 'TranslationJobError' in message and 'unregistered' not in message:
                job['approved_error'] = timestamp
                job['status'] = 'ERROR'

    return jobs, thread_to_jobs


def parse_api_errors(logfile, date_str):
    """Parse API call errors for retry detection. Returns list of dicts."""
    errors = []
    tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False)
    tmp.close()

    for pattern in ['Error for Azure Foundry API call', 'Error processing access token request']:
        grep_extract(logfile, pattern, tmp.name)
        with open(tmp.name, 'r') as f:
            for line in f:
                m = LINE_RE.match(line.strip())
                if m:
                    _date, timestamp, _inst, _level, thread, _cls, message = m.groups()
                    wf_match = WORKFLOW_RE.search(thread)
                    workflow_prefix = wf_match.group(1) if wf_match else thread
                    etype = 'api_payload_dump' if 'Azure Foundry' in message else 'token_request_error'
                    errors.append({
                        'timestamp': timestamp,
                        'thread': thread,
                        'workflow_prefix': workflow_prefix,
                        'type': etype,
                    })

    os.unlink(tmp.name)
    return errors


def detect_retries(jobs, thread_to_jobs, api_errors):
    """Correlate API errors with specific jobs using timing and workflow thread.
    thread_to_jobs maps workflow_prefix -> set of composite keys (project/jobId).
    """
    workflow_errors = defaultdict(list)
    for err in api_errors:
        workflow_errors[err['workflow_prefix']].append(err)

    for wf_prefix, error_list in workflow_errors.items():
        if wf_prefix not in thread_to_jobs:
            continue

        wf_jobs = [jobs[ck] for ck in thread_to_jobs[wf_prefix] if ck in jobs]

        for err in error_list:
            err_ms = parse_time(err['timestamp'])
            if err_ms is None:
                continue

            best_job = None
            best_diff = float('inf')

            for job in wf_jobs:
                ca_end = parse_time(job['content_add_end']) if job['content_add_end'] else None
                created_ms = parse_time(job['created']) if job['created'] else None
                start_ms = ca_end or created_ms
                if start_ms is None:
                    continue

                end_ms = parse_time(job['approved_error']) if job['approved_error'] else None
                ip_ms = parse_time(job['in_progress']) if job['in_progress'] else None
                final_ms = end_ms or ip_ms

                if start_ms <= err_ms:
                    if final_ms is None or err_ms <= final_ms:
                        diff = err_ms - start_ms
                        if diff < best_diff:
                            best_diff = diff
                            best_job = job

            if best_job:
                existing = int(best_job.get('retries', '0') or '0')
                best_job['retries'] = str(existing + 1)
                if best_job['status'] == 'APPROVED':
                    best_job['retry_detail'] = f"API call failed ({err['type']}) → retried → APPROVED"
                else:
                    best_job['retry_detail'] = f"API call failed ({err['type']})"

# ---------------------------------------------------------------------------
# RMTranslationRequest parsing
# ---------------------------------------------------------------------------

def parse_rm_translation_requests(logfile):
    """Parse RMTranslationRequest log entries. Returns summary dict."""
    tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False)
    tmp.close()

    count = grep_extract(logfile, 'RMTranslationRequest', tmp.name)
    summary = {
        'total': count,
        'init': 0,
        'obtained': 0,
        'expires': 0,
        'errors': 0,
        'token_size': 0,
    }

    with open(tmp.name, 'r') as f:
        for line in f:
            if 'initialized successfully' in line:
                summary['init'] += 1
            elif 'Successfully obtained access token' in line:
                summary['obtained'] += 1
            elif 'Token expires at' in line or 'access token expires' in line:
                summary['expires'] += 1
            elif 'Error processing access token request' in line:
                summary['errors'] += 1
            elif 'token size of chatRequestPromptJson' in line:
                summary['token_size'] += 1

    os.unlink(tmp.name)
    return summary

# ---------------------------------------------------------------------------
# Error categorization
# ---------------------------------------------------------------------------

def categorize_errors(logfile):
    """Count errors by category. Returns OrderedDict of category->count."""
    counts = OrderedDict()
    for pattern, label in ERROR_CATEGORIES.items():
        c = grep_count(logfile, pattern)
        if c > 0:
            counts[label] = c

    noise = OrderedDict()
    for pat in NOISE_PATTERNS:
        c = grep_count(logfile, pat)
        if c > 0:
            noise[pat] = c

    return counts, noise

# ---------------------------------------------------------------------------
# CSV operations
# ---------------------------------------------------------------------------

def load_existing_csv(csv_path):
    """Load existing CSV data, returns list of dicts."""
    if not os.path.exists(csv_path):
        return []
    rows = []
    with open(csv_path, 'r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def jobs_to_rows(jobs, date_str):
    """Convert parsed jobs dict to CSV row dicts."""
    sorted_jobs = sorted(jobs.values(), key=lambda j: j.get('created', '') or 'zz')
    rows = []
    for idx, job in enumerate(sorted_jobs, 1):
        cg_ms = diff_ms(job['content_add_start'], job['content_add_end']) if job['content_add_start'] and job['content_add_end'] else None
        ci_ms = diff_ms(job['created'], job['in_progress']) if job['created'] and job['in_progress'] else None
        ir_ms = diff_ms(job['in_progress'], job['ready_for_review']) if job['in_progress'] and job['ready_for_review'] else None
        ra_ms = diff_ms(job['ready_for_review'], job['approved_error']) if job['ready_for_review'] and job['approved_error'] else None
        total_ms = diff_ms(job['created'], job['approved_error']) if job['created'] and job['approved_error'] else None

        row = {
            'Date': date_str,
            '#': str(idx),
            'Job ID': job['job_id'],
            'Project': job['project'],
            'Language': job['language'],
            'Source Language': job['source_language'],
            'Connector': job['connector'],
            'Method': job['method'],
            'Status': job['status'],
            'Created': job['created'],
            'Content Add Start': job['content_add_start'],
            'Content Add End': job['content_add_end'],
            'InProgress': job['in_progress'],
            'ReadyForReview': job['ready_for_review'],
            'Approved/Error': job['approved_error'],
            'Content Gathering (ms)': str(cg_ms) if cg_ms is not None else '',
            'Content Gathering': ms_to_human(cg_ms) if cg_ms is not None else '',
            'Created→InProg (ms)': str(ci_ms) if ci_ms is not None else '',
            'Created→InProg': ms_to_human(ci_ms) if ci_ms is not None else '',
            'InProg→Ready (ms)': str(ir_ms) if ir_ms is not None else '',
            'InProg→Ready': ms_to_human(ir_ms) if ir_ms is not None else '',
            'Ready→Approved (ms)': str(ra_ms) if ra_ms is not None else '',
            'Ready→Approved': ms_to_human(ra_ms) if ra_ms is not None else '',
            'Total Duration (ms)': str(total_ms) if total_ms is not None else '',
            'Total Duration': ms_to_human(total_ms) if total_ms is not None else '',
            'Content Objects': job['content_objects'],
            'Content Size': job['content_size'],
            'Errors': job['errors'],
            'Retries': job.get('retries', ''),
            'Retry Detail': job.get('retry_detail', ''),
        }
        rows.append(row)
    return rows


def save_csv(rows, csv_path):
    """Save rows to CSV."""
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------

HTML_CSS = """
:root{--bg:#0f1117;--surface:#1a1d27;--surface2:#232733;--border:#2e3345;--text:#e2e4ea;--muted:#8b8fa3;
--accent:#6c8aff;--green:#34d399;--green-dim:rgba(52,211,153,.12);--red:#f87171;--red-dim:rgba(248,113,113,.12);
--orange:#fbbf24;--orange-dim:rgba(251,191,36,.12);--yellow:#fde047;--yellow-dim:rgba(253,224,71,.12);
--blue:#60a5fa;--blue-dim:rgba(96,165,250,.12);--purple:#a78bfa;--purple-dim:rgba(167,139,250,.12);
--cyan:#22d3ee;--cyan-dim:rgba(34,211,238,.12);--pink:#f472b6;--pink-dim:rgba(244,114,182,.12)}
*{margin:0;padding:0;box-sizing:border-box}
body{background:var(--bg);color:var(--text);font-family:'Inter','Segoe UI',system-ui,sans-serif;font-size:13px;line-height:1.5;padding:24px}
.ctr{max-width:1700px;margin:0 auto}h1{font-size:22px;font-weight:700;margin-bottom:4px}
h2{font-size:17px;font-weight:600;margin:24px 0 10px;color:var(--accent)}h3{font-size:14px;font-weight:600;margin:16px 0 8px}
.sub{color:var(--muted);font-size:13px;margin-bottom:20px}
.sr{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:20px}
.sc{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;min-width:140px}
.sl{color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.5px}
.sv{font-size:22px;font-weight:700;margin-top:2px}.sv.g{color:var(--green)}.sv.o{color:var(--orange)}
.sv.r{color:var(--red)}.sv.a{color:var(--accent)}.sv.p{color:var(--purple)}.sv.c{color:var(--cyan)}.sd{color:var(--muted);font-size:11px}
.tw{overflow-x:auto;border:1px solid var(--border);border-radius:10px;margin-bottom:20px}
table{width:100%;border-collapse:collapse;font-size:12px;white-space:nowrap}
th{background:var(--surface2);color:var(--muted);font-weight:600;font-size:10px;text-transform:uppercase;letter-spacing:.4px;padding:10px;text-align:left;border-bottom:1px solid var(--border);position:sticky;top:0;z-index:1}
td{padding:8px 10px;border-bottom:1px solid var(--border)}tr:hover td{background:rgba(108,138,255,.04)}
tr:last-child td{border-bottom:none}.m{font-family:'JetBrains Mono','Fira Code',monospace;font-size:11px}.tr{text-align:right}
.status-ok{background:var(--green-dim);color:var(--green);font-weight:600;padding:2px 8px;border-radius:4px;display:inline-block;font-size:10px}
.status-warn{background:var(--orange-dim);color:var(--orange);font-weight:600;padding:2px 8px;border-radius:4px;display:inline-block;font-size:10px}
.status-err{background:var(--red-dim);color:var(--red);font-weight:600;padding:2px 8px;border-radius:4px;display:inline-block;font-size:10px}
.status-unk{background:var(--yellow-dim);color:var(--yellow);font-weight:600;padding:2px 8px;border-radius:4px;display:inline-block;font-size:10px}
.row-warn td{background:var(--orange-dim)}.row-err td{background:var(--red-dim)}.row-unk td{background:var(--yellow-dim)}
.row-retry td{background:var(--cyan-dim)}
.retry-badge{background:var(--cyan-dim);color:var(--cyan);font-weight:600;padding:2px 8px;border-radius:4px;display:inline-block;font-size:10px}
.db{display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:600;margin-right:4px}
.eg{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:12px;margin-bottom:20px}
.ec{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px}
.ec h4{font-size:12px;margin-bottom:6px}.ec .cnt{font-size:18px;font-weight:700}
.ec .cnt.red{color:var(--red)}.ec .cnt.orange{color:var(--orange)}.ec .cnt.muted{color:var(--muted)}
.notice{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px;margin-bottom:20px}
.notice h4{margin-bottom:6px}
.footer{color:var(--muted);text-align:center;font-size:11px;margin-top:24px}
"""

DATE_COLORS = [
    ('var(--blue)', 'var(--blue-dim)'),
    ('var(--green)', 'var(--green-dim)'),
    ('var(--orange)', 'var(--orange-dim)'),
    ('var(--purple)', 'var(--purple-dim)'),
    ('var(--pink)', 'var(--pink-dim)'),
    ('var(--cyan)', 'var(--cyan-dim)'),
    ('var(--yellow)', 'var(--yellow-dim)'),
    ('var(--red)', 'var(--red-dim)'),
]


def status_class(s):
    s = s.upper()
    if s == 'APPROVED':
        return 'status-ok'
    if s == 'ERROR':
        return 'status-err'
    if s in ('IN_PROGRESS', 'READY_FOR_REVIEW'):
        return 'status-warn'
    return 'status-unk'


def generate_html(all_rows, errors_by_date, rm_tokens_by_date, html_path):
    """Generate consolidated HTML report."""
    dates = sorted(set(r['Date'] for r in all_rows))
    date_color_map = {d: DATE_COLORS[i % len(DATE_COLORS)] for i, d in enumerate(dates)}

    total_jobs = len(all_rows)
    approved = sum(1 for r in all_rows if r['Status'] == 'APPROVED')
    stuck = sum(1 for r in all_rows if r['Status'] in ('IN_PROGRESS', 'READY_FOR_REVIEW'))
    unknown = sum(1 for r in all_rows if r['Status'] == 'UNKNOWN')
    errored = sum(1 for r in all_rows if r['Status'] == 'ERROR')
    retried = sum(1 for r in all_rows if r.get('Retries') and int(r.get('Retries', '0') or '0') > 0)

    durations = []
    for r in all_rows:
        v = r.get('Total Duration (ms)', '')
        if v:
            try:
                durations.append(int(v))
            except ValueError:
                pass
    avg_dur = sum(durations) / len(durations) if durations else 0
    min_dur = min(durations) if durations else 0
    max_dur = max(durations) if durations else 0

    date_badges = ' '.join(
        f'<span class="db" style="background:{date_color_map[d][1]};color:{date_color_map[d][0]}">{format_date_display(d)}</span>'
        for d in dates
    )

    html = f'''<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Translation Jobs Report - PROD Author ({len(dates)} day{"s" if len(dates)!=1 else ""})</title>
<style>{HTML_CSS}</style></head><body><div class="ctr">

<h1>Translation Jobs Report &mdash; PROD Author ({len(dates)} Day{"s" if len(dates)!=1 else ""})</h1>
<p class="sub">Logs: {date_badges} &nbsp;|&nbsp; <span class="retry-badge">RETRY TRACKING</span></p>

<div class="sr">
<div class="sc"><div class="sl">Total Jobs</div><div class="sv a">{total_jobs}</div></div>
<div class="sc"><div class="sl">Approved</div><div class="sv g">{approved}</div></div>
<div class="sc"><div class="sl">Errors</div><div class="sv r">{errored}</div></div>
<div class="sc"><div class="sl">Stuck</div><div class="sv o">{stuck}</div></div>
<div class="sc"><div class="sl">Unknown</div><div class="sv o">{unknown}</div></div>
<div class="sc"><div class="sl">Retried &amp; Recovered</div><div class="sv c">{retried}</div></div>
<div class="sc"><div class="sl">Avg Duration</div><div class="sv a">{ms_to_human(avg_dur)}</div>
<div class="sd">Min {ms_to_human(min_dur)} / Max {ms_to_human(max_dur)}</div></div>
</div>
'''

    # Retry detection summary
    retry_rows = [r for r in all_rows if r.get('Retries') and int(r.get('Retries', '0') or '0') > 0]
    if retry_rows:
        html += '<div class="notice" style="border-color:var(--cyan)"><h4 style="color:var(--cyan)">Retry Detection</h4>\n'
        html += '<p style="color:var(--muted);font-size:12px">Detected by correlating <code>Error for Azure Foundry API call</code> / <code>Error processing access token request</code> with translation pipeline events.</p>\n'
        html += '<table style="width:auto;font-size:12px;margin-top:8px"><thead><tr><th>Date</th><th>#</th><th>Job ID</th><th>Project</th><th>Language</th><th>Detail</th><th>Status</th></tr></thead><tbody>\n'
        for r in retry_rows:
            col, bg = date_color_map.get(r['Date'], DATE_COLORS[0])
            st_cls = status_class(r['Status'])
            html += f'<tr><td><span class="db" style="background:{bg};color:{col}">{r["Date"]}</span></td>'
            html += f'<td class="m">{r["#"]}</td><td class="m">{r["Job ID"]}</td>'
            html += f'<td>{r["Project"]}</td><td class="m">{r["Language"]}</td>'
            html += f'<td style="color:var(--cyan);font-size:10px">{r.get("Retry Detail","")}</td>'
            html += f'<td><span class="{st_cls}">{r["Status"]}</span></td></tr>\n'
        html += '</tbody></table></div>\n'

    # Per-day summaries
    for d in dates:
        day_rows = [r for r in all_rows if r['Date'] == d]
        col, bg = date_color_map[d]
        d_approved = sum(1 for r in day_rows if r['Status'] == 'APPROVED')
        d_stuck = sum(1 for r in day_rows if r['Status'] in ('IN_PROGRESS', 'READY_FOR_REVIEW'))
        d_unk = sum(1 for r in day_rows if r['Status'] == 'UNKNOWN')
        d_err = sum(1 for r in day_rows if r['Status'] == 'ERROR')
        d_retry = sum(1 for r in day_rows if r.get('Retries') and int(r.get('Retries', '0') or '0') > 0)

        html += f'<h2><span class="db" style="background:{bg};color:{col}">{d}</span> &mdash; {len(day_rows)} jobs</h2>\n'
        html += f'<div class="sr">'
        html += f'<div class="sc"><div class="sl">Approved</div><div class="sv g">{d_approved}</div></div>'
        if d_err:
            html += f'<div class="sc"><div class="sl">Errors</div><div class="sv r">{d_err}</div></div>'
        if d_stuck:
            html += f'<div class="sc"><div class="sl">Stuck</div><div class="sv o">{d_stuck}</div></div>'
        if d_unk:
            html += f'<div class="sc"><div class="sl">Unknown</div><div class="sv o">{d_unk}</div></div>'
        if d_retry:
            html += f'<div class="sc"><div class="sl">Retried</div><div class="sv c">{d_retry}</div></div>'
        html += '</div>\n'

        # Error summary for this date
        if d in errors_by_date and errors_by_date[d]:
            html += '<div class="eg">\n'
            for cat, cnt in errors_by_date[d].items():
                color_cls = 'red' if cnt > 100 else ('orange' if cnt > 10 else 'muted')
                html += f'<div class="ec"><h4>{cat}</h4><div class="cnt {color_cls}">{cnt:,}</div></div>\n'
            html += '</div>\n'

        # RMTranslationRequest summary
        if d in rm_tokens_by_date:
            tk = rm_tokens_by_date[d]
            html += f'<div class="notice"><h4>RMTranslationRequest Activity</h4>'
            html += f'<p style="color:var(--muted);font-size:12px">{tk["total"]} entries: {tk["init"]} init + {tk["obtained"]} obtained + {tk["expires"]} expires'
            if tk['errors']:
                html += f' + <span style="color:var(--red)">{tk["errors"]} errors</span>'
            if tk['token_size']:
                html += f' + {tk["token_size"]} token size entries'
            else:
                html += f' + 0 chatRequestPromptJson (DEBUG level needed)'
            html += '</p></div>\n'

    # Full job table
    html += '<h2>All Translation Jobs</h2>\n<div class="tw"><table>\n<thead><tr>'
    display_cols = ['Date', '#', 'Job ID', 'Project', 'Language', 'Status',
                    'Created', 'InProgress', 'ReadyForReview', 'Approved/Error',
                    'Content Gathering', 'Created→InProg', 'InProg→Ready',
                    'Ready→Approved', 'Total Duration', 'Content Objects', 'Retries']
    for c in display_cols:
        html += f'<th>{c}</th>'
    html += '</tr></thead>\n<tbody>\n'

    for r in all_rows:
        row_class = ''
        if r.get('Retries') and int(r.get('Retries', '0') or '0') > 0:
            row_class = ' class="row-retry"'
        elif r['Status'] == 'ERROR':
            row_class = ' class="row-err"'
        elif r['Status'] in ('IN_PROGRESS', 'READY_FOR_REVIEW'):
            row_class = ' class="row-warn"'
        elif r['Status'] == 'UNKNOWN':
            row_class = ' class="row-unk"'

        col, bg = date_color_map.get(r['Date'], DATE_COLORS[0])
        st_cls = status_class(r['Status'])

        html += f'<tr{row_class}>'
        html += f'<td><span class="db" style="background:{bg};color:{col}">{format_date_display(r["Date"])}</span></td>'
        html += f'<td class="m">{r["#"]}</td>'
        html += f'<td class="m">{r["Job ID"]}</td>'
        html += f'<td>{r["Project"]}</td>'
        html += f'<td class="m">{r["Language"]}</td>'
        html += f'<td><span class="{st_cls}">{r["Status"]}</span></td>'
        html += f'<td class="m">{r["Created"]}</td>'
        html += f'<td class="m">{r["InProgress"]}</td>'
        html += f'<td class="m">{r["ReadyForReview"]}</td>'
        html += f'<td class="m">{r["Approved/Error"]}</td>'
        html += f'<td class="m tr">{r.get("Content Gathering","")}</td>'
        html += f'<td class="m tr">{r.get("Created→InProg","")}</td>'
        html += f'<td class="m tr">{r.get("InProg→Ready","")}</td>'
        html += f'<td class="m tr">{r.get("Ready→Approved","")}</td>'
        html += f'<td class="m tr">{r.get("Total Duration","")}</td>'
        html += f'<td class="m" style="font-size:9px">{r.get("Content Objects","")}</td>'

        retries_val = r.get('Retries', '')
        if retries_val and int(retries_val or '0') > 0:
            html += f'<td><span class="retry-badge">{retries_val}</span></td>'
        else:
            html += '<td></td>'
        html += '</tr>\n'

    html += '</tbody></table></div>\n'

    html += f'<div class="footer">Generated {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} by AEM Translation Log Analyzer</div>\n'
    html += '</div></body></html>'

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)

# ---------------------------------------------------------------------------
# XLSX report
# ---------------------------------------------------------------------------

def generate_xlsx(all_rows, errors_by_date, rm_tokens_by_date, xlsx_path):
    """Generate XLSX workbook with multiple sheets."""
    if not HAS_OPENPYXL:
        print('  Warning: openpyxl not installed, skipping XLSX generation', file=sys.stderr)
        return

    wb = Workbook()

    # --- Sheet 1: Translation Jobs ---
    ws = wb.active
    ws.title = 'Translation Jobs'

    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='2E3345', end_color='2E3345', fill_type='solid')
    ok_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    err_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    warn_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    retry_fill = PatternFill(start_color='CFFAFE', end_color='CFFAFE', fill_type='solid')
    mono_font = Font(name='Consolas', size=9)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    for ci, col_name in enumerate(CSV_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for ri, row in enumerate(all_rows, 2):
        for ci, col_name in enumerate(CSV_COLUMNS, 1):
            val = row.get(col_name, '')
            if col_name in ('Content Gathering (ms)', 'Created→InProg (ms)', 'InProg→Ready (ms)',
                            'Ready→Approved (ms)', 'Total Duration (ms)', '#', 'Content Size', 'Retries'):
                try:
                    val = int(val) if val else ''
                except ValueError:
                    pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = mono_font
            cell.border = thin_border

        status = row.get('Status', '')
        retries = row.get('Retries', '')
        fill = None
        if retries and int(retries or '0') > 0:
            fill = retry_fill
        elif status == 'APPROVED':
            fill = ok_fill
        elif status == 'ERROR':
            fill = err_fill
        elif status in ('UNKNOWN', 'IN_PROGRESS', 'READY_FOR_REVIEW'):
            fill = warn_fill
        if fill:
            for ci in range(1, len(CSV_COLUMNS) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    for ci in range(1, len(CSV_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(CSV_COLUMNS[ci - 1]) + 4)

    ws.auto_filter.ref = f'A1:{get_column_letter(len(CSV_COLUMNS))}{len(all_rows) + 1}'
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    dates = sorted(set(r['Date'] for r in all_rows))
    summary_headers = ['Date', 'Total Jobs', 'Approved', 'Errors', 'Stuck', 'Unknown', 'Retried',
                       'Avg Duration', 'Min Duration', 'Max Duration']
    for ci, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for ri, d in enumerate(dates, 2):
        day_rows = [r for r in all_rows if r['Date'] == d]
        durs = []
        for r in day_rows:
            v = r.get('Total Duration (ms)', '')
            if v:
                try:
                    durs.append(int(v))
                except ValueError:
                    pass
        vals = [
            d,
            len(day_rows),
            sum(1 for r in day_rows if r['Status'] == 'APPROVED'),
            sum(1 for r in day_rows if r['Status'] == 'ERROR'),
            sum(1 for r in day_rows if r['Status'] in ('IN_PROGRESS', 'READY_FOR_REVIEW')),
            sum(1 for r in day_rows if r['Status'] == 'UNKNOWN'),
            sum(1 for r in day_rows if r.get('Retries') and int(r.get('Retries', '0') or '0') > 0),
            ms_to_human(sum(durs) / len(durs)) if durs else '',
            ms_to_human(min(durs)) if durs else '',
            ms_to_human(max(durs)) if durs else '',
        ]
        for ci, v in enumerate(vals, 1):
            ws2.cell(row=ri, column=ci, value=v).border = thin_border

    for ci in range(1, len(summary_headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 16

    # --- Sheet 3: Errors ---
    ws3 = wb.create_sheet('Errors by Date')
    all_cats = list(OrderedDict.fromkeys(
        cat for d in dates if d in errors_by_date for cat in errors_by_date[d]
    ))
    err_headers = ['Error Category'] + dates
    for ci, h in enumerate(err_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for ri, cat in enumerate(all_cats, 2):
        ws3.cell(row=ri, column=1, value=cat).border = thin_border
        for di, d in enumerate(dates, 2):
            v = errors_by_date.get(d, {}).get(cat, 0)
            cell = ws3.cell(row=ri, column=di, value=v if v else '')
            cell.border = thin_border
            if v and v > 100:
                cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for ci in range(1, len(err_headers) + 1):
        ws3.column_dimensions[get_column_letter(ci)].width = 30 if ci == 1 else 14

    # --- Sheet 4: RMTranslationRequest ---
    ws4 = wb.create_sheet('RMTranslationRequest')
    tk_headers = ['Date', 'Total Entries', 'Init', 'Obtained', 'Expires', 'Errors', 'Token Size Entries']
    for ci, h in enumerate(tk_headers, 1):
        cell = ws4.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for ri, d in enumerate(dates, 2):
        tk = rm_tokens_by_date.get(d, {'total': 0, 'init': 0, 'obtained': 0, 'expires': 0, 'errors': 0, 'token_size': 0})
        vals = [d, tk['total'], tk['init'], tk['obtained'], tk['expires'], tk['errors'], tk['token_size']]
        for ci, v in enumerate(vals, 1):
            cell = ws4.cell(row=ri, column=ci, value=v)
            cell.border = thin_border
            if ci == 6 and v > 0:
                cell.fill = err_fill

    for ci in range(1, len(tk_headers) + 1):
        ws4.column_dimensions[get_column_letter(ci)].width = 18

    wb.save(xlsx_path)

# ---------------------------------------------------------------------------
# Metadata persistence (errors, tokens per date)
# ---------------------------------------------------------------------------

def load_metadata(meta_path):
    """Load per-date errors and token metadata from JSON."""
    if os.path.exists(meta_path):
        with open(meta_path, 'r') as f:
            return json.load(f)
    return {'errors': {}, 'tokens': {}, 'noise': {}}


def save_metadata(meta, meta_path):
    with open(meta_path, 'w') as f:
        json.dump(meta, f, indent=2)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='AEM Translation Log Analyzer')
    parser.add_argument('logfile', help='Path to aemerror log file')
    parser.add_argument('--reports-dir', default='reports', help='Output directory for reports')
    parser.add_argument('--date', help='Override date (YYYY-MM-DD), otherwise detected from filename')
    args = parser.parse_args()

    logfile = os.path.abspath(args.logfile)
    reports_dir = os.path.abspath(args.reports_dir)
    os.makedirs(reports_dir, exist_ok=True)

    date_str = args.date or extract_date_from_filename(logfile)
    if not date_str:
        print('Error: could not detect date from filename. Use --date YYYY-MM-DD', file=sys.stderr)
        sys.exit(1)

    print(f'Analyzing {os.path.basename(logfile)} for date {date_str}')
    file_size_mb = os.path.getsize(logfile) / (1024 * 1024)
    print(f'  File size: {file_size_mb:.0f} MB')

    # Step 1: Extract translation pipeline events
    print('  Extracting translation pipeline events...')
    events_tmp = tempfile.NamedTemporaryFile(mode='w', suffix='_events.txt', delete=False)
    events_tmp.close()
    ev_count = grep_extract(logfile, 'Translation pipeline Event', events_tmp.name)
    print(f'    Found {ev_count} event lines')

    # Step 2: Parse events into job records
    print('  Parsing events...')
    jobs, thread_to_jobs = parse_events(events_tmp.name, date_str)
    os.unlink(events_tmp.name)
    print(f'    Found {len(jobs)} translation jobs')

    # Step 3: Detect retries
    print('  Checking for API errors (retry detection)...')
    api_errors = parse_api_errors(logfile, date_str)
    print(f'    Found {len(api_errors)} API errors')
    if api_errors:
        detect_retries(jobs, thread_to_jobs, api_errors)
        retried = sum(1 for j in jobs.values() if j.get('retries'))
        print(f'    Detected {retried} retried jobs')

    # Step 4: Parse RMTranslationRequest entries
    print('  Parsing RMTranslationRequest entries...')
    rm_tokens = parse_rm_translation_requests(logfile)
    print(f'    {rm_tokens["total"]} total: {rm_tokens["init"]} init, {rm_tokens["obtained"]} obtained, {rm_tokens["errors"]} errors')

    # Step 5: Categorize errors
    print('  Categorizing errors...')
    error_counts, noise_counts = categorize_errors(logfile)
    for cat, cnt in error_counts.items():
        print(f'    {cat}: {cnt:,}')
    if noise_counts:
        print(f'    Noise filtered: {sum(noise_counts.values()):,} entries')

    # Step 6: Load existing data and merge
    csv_path = os.path.join(reports_dir, 'translation_jobs_all_days.csv')
    html_path = os.path.join(reports_dir, 'translation_jobs_all_days.html')
    xlsx_path = os.path.join(reports_dir, 'translation_jobs_all_days.xlsx')
    meta_path = os.path.join(reports_dir, 'metadata.json')

    existing_rows = load_existing_csv(csv_path)
    filtered_rows = [r for r in existing_rows if r.get('Date') != date_str]
    print(f'  Existing data: {len(existing_rows)} rows ({len(existing_rows) - len(filtered_rows)} for {date_str} will be replaced)')

    new_rows = jobs_to_rows(jobs, date_str)
    all_rows = filtered_rows + new_rows
    all_rows.sort(key=lambda r: (r['Date'], int(r.get('#', '0') or '0')))

    # Renumber within each date
    date_counters = defaultdict(int)
    for r in all_rows:
        date_counters[r['Date']] += 1
        r['#'] = str(date_counters[r['Date']])

    # Step 7: Load/update metadata
    meta = load_metadata(meta_path)
    meta['errors'][date_str] = dict(error_counts)
    meta['tokens'][date_str] = rm_tokens
    meta['noise'][date_str] = dict(noise_counts)
    save_metadata(meta, meta_path)

    errors_by_date = {d: OrderedDict(v) for d, v in meta['errors'].items()}
    rm_tokens_by_date = meta['tokens']

    # Step 8: Generate reports
    print(f'  Generating CSV ({len(all_rows)} rows)...')
    save_csv(all_rows, csv_path)

    print('  Generating HTML...')
    generate_html(all_rows, errors_by_date, rm_tokens_by_date, html_path)

    print('  Generating XLSX...')
    generate_xlsx(all_rows, errors_by_date, rm_tokens_by_date, xlsx_path)

    print(f'\nDone! Reports saved to {reports_dir}/')
    print(f'  - {os.path.basename(csv_path)}')
    print(f'  - {os.path.basename(html_path)}')
    print(f'  - {os.path.basename(xlsx_path)}')

    # Summary
    dates = sorted(set(r['Date'] for r in all_rows))
    print(f'\nConsolidated report covers {len(dates)} day(s): {", ".join(dates)}')
    print(f'  Total jobs: {len(all_rows)}')
    print(f'  Approved: {sum(1 for r in all_rows if r["Status"] == "APPROVED")}')
    print(f'  Errors: {sum(1 for r in all_rows if r["Status"] == "ERROR")}')
    print(f'  Retried: {sum(1 for r in all_rows if r.get("Retries") and int(r.get("Retries","0") or "0") > 0)}')


if __name__ == '__main__':
    main()
