#!/usr/bin/env python3
"""AEM PROD Publish aemerror log analyzer - general error validation."""

import sys
import os
import re
import subprocess
import argparse
from datetime import datetime
from collections import defaultdict, OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

LINE_RE = re.compile(
    r'^(\d{2}\.\d{2}\.\d{4})\s+'          # date DD.MM.YYYY
    r'(\d{2}:\d{2}:\d{2}\.\d{3})\s+'      # timestamp
    r'\[([^\]]+)\]\s+\*(\w+)\*\s+\[([^\]]+)\]\s+(\S+)\s+(.*)'
)

def grep_extract(logfile, pattern, outfile):
    try:
        r = subprocess.run(['grep', pattern, logfile], capture_output=True, text=True, timeout=600)
        with open(outfile, 'w') as f:
            f.write(r.stdout)
        return len(r.stdout)
    except Exception:
        return 0

def parse_errors(logfile):
    """Extract ERROR and WARN lines, categorize by class."""
    errs, warns = [], []
    tmp_err = '/tmp/publish_err.txt'
    tmp_warn = '/tmp/publish_warn.txt'
    grep_extract(logfile, r'\*ERROR\*', tmp_err)
    with open(tmp_err) as f:
        for line in f:
            m = LINE_RE.match(line.strip())
            if m:
                errs.append({'date': m.group(1), 'time': m.group(2), 'level': 'ERROR', 'class': m.group(6), 'msg': m.group(7)[:200]})
    grep_extract(logfile, r'\*WARN\*', tmp_warn)
    with open(tmp_warn) as f:
        for line in f:
            m = LINE_RE.match(line.strip())
            if m:
                warns.append({'date': m.group(1), 'time': m.group(2), 'level': 'WARN', 'class': m.group(6), 'msg': m.group(7)[:200]})
    return errs, warns

def generate_html(errs, warns, by_class, out_path):
    html = '''<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Publish Error Report</title>
<style>
body{font-family:system-ui,sans-serif;margin:2rem;background:#1a1a2e;color:#eee;}
h1,h2{color:#e94560;}
table{border-collapse:collapse;width:100%;margin:1rem 0;}
th,td{border:1px solid #444;padding:8px;text-align:left;}
th{background:#16213e;color:#e94560;}
tr:nth-child(even){background:#16213e;}
.error{color:#e94560;}
.warn{color:#f0ad4e;}
a{color:#4fc3f7;}
</style></head><body>
<h1>AEM PROD Publish Error Report</h1>
<p>Generated from aemerror.log — general error validation. <a href="../">← Translation Report</a></p>
<h2>Summary</h2>
<table><tr><th>Level</th><th>Count</th></tr>
<tr><td class="error">ERROR</td><td>''' + str(len(errs)) + '''</td></tr>
<tr><td class="warn">WARN</td><td>''' + str(len(warns)) + '''</td></tr>
</table>
<h2>By Class</h2>
<table><tr><th>Class</th><th>ERROR</th><th>WARN</th></tr>'''
    for cls, counts in sorted(by_class.items(), key=lambda x: -(x[1].get('ERROR', 0) + x[1].get('WARN', 0))):
        html += f'<tr><td>{cls}</td><td>{counts.get("ERROR",0)}</td><td>{counts.get("WARN",0)}</td></tr>'
    html += '''</table>
<h2>Recent Errors (sample)</h2>
<table><tr><th>Date</th><th>Time</th><th>Class</th><th>Message</th></tr>'''
    for r in errs[:100]:
        html += f'<tr><td>{r["date"]}</td><td>{r["time"]}</td><td>{r["class"]}</td><td>{r["msg"][:150]}</td></tr>'
    html += '''</table>
<h2>Recent Warnings (sample)</h2>
<table><tr><th>Date</th><th>Time</th><th>Class</th><th>Message</th></tr>'''
    for r in warns[:100]:
        html += f'<tr><td>{r["date"]}</td><td>{r["time"]}</td><td>{r["class"]}</td><td>{r["msg"][:150]}</td></tr>'
    html += '</table></body></html>'
    with open(out_path, 'w') as f:
        f.write(html)

def generate_md(errs, warns, by_class, out_path):
    lines = ['# AEM PROD Publish Error Report\n', 'Generated from aemerror.log — general error validation.\n',
              '## Summary\n', f'- **ERROR**: {len(errs)}', f'- **WARN**: {len(warns)}\n',
              '## By Class\n', '| Class | ERROR | WARN |\n|-------|-------|------|\n']
    for cls, counts in sorted(by_class.items(), key=lambda x: -(x[1].get('ERROR', 0) + x[1].get('WARN', 0))):
        lines.append(f'| {cls} | {counts.get("ERROR",0)} | {counts.get("WARN",0)} |\n')
    lines.append('\n## Recent Errors (sample)\n| Date | Time | Class | Message |\n|------|------|-------|--------|\n')
    for r in errs[:50]:
        lines.append(f'| {r["date"]} | {r["time"]} | {r["class"]} | {r["msg"][:100]} |\n')
    lines.append('\n## Recent Warnings (sample)\n| Date | Time | Class | Message |\n|------|------|-------|--------|\n')
    for r in warns[:50]:
        lines.append(f'| {r["date"]} | {r["time"]} | {r["class"]} | {r["msg"][:100]} |\n')
    with open(out_path, 'w') as f:
        f.writelines(lines)

def generate_xlsx(errs, warns, by_class, out_path):
    if not HAS_OPENPYXL:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    ws.append(['Level', 'Count'])
    ws.append(['ERROR', len(errs)])
    ws.append(['WARN', len(warns)])
    ws2 = wb.create_sheet('By Class')
    ws2.append(['Class', 'ERROR', 'WARN'])
    for cls, counts in sorted(by_class.items(), key=lambda x: -(x[1].get('ERROR', 0) + x[1].get('WARN', 0))):
        ws2.append([cls, counts.get('ERROR', 0), counts.get('WARN', 0)])
    ws3 = wb.create_sheet('Errors')
    ws3.append(['Date', 'Time', 'Class', 'Message'])
    for r in errs:
        ws3.append([r['date'], r['time'], r['class'], r['msg']])
    ws4 = wb.create_sheet('Warnings')
    ws4.append(['Date', 'Time', 'Class', 'Message'])
    for r in warns:
        ws4.append([r['date'], r['time'], r['class'], r['msg']])
    wb.save(out_path)

def main():
    parser = argparse.ArgumentParser(description='Analyze AEM publish aemerror log')
    parser.add_argument('logfile', help='Path to aemerror log')
    parser.add_argument('--reports-dir', default='reports/publish', help='Output directory')
    args = parser.parse_args()
    if not os.path.isfile(args.logfile):
        print(f'Error: log file not found: {args.logfile}')
        sys.exit(1)
    os.makedirs(args.reports_dir, exist_ok=True)
    errs, warns = parse_errors(args.logfile)
    by_class = defaultdict(lambda: {'ERROR': 0, 'WARN': 0})
    for r in errs:
        by_class[r['class']]['ERROR'] += 1
    for r in warns:
        by_class[r['class']]['WARN'] += 1
    generate_html(errs, warns, by_class, os.path.join(args.reports_dir, 'index.html'))
    generate_md(errs, warns, by_class, os.path.join(args.reports_dir, 'SUMMARY.md'))
    generate_xlsx(errs, warns, by_class, os.path.join(args.reports_dir, 'errors.xlsx'))
    print(f'Publish report: {len(errs)} errors, {len(warns)} warnings')

if __name__ == '__main__':
    main()
